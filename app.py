import streamlit as st
import pandas as pd
import os

from database import (
    get_supabase,
    fetch_table,
    insert_data,
    upsert_data,
    get_config_val,
    get_origem_perguntas,
    get_perguntas_por_origem,
)
from pricing import calcular_preco_completo
from validators import (
    validar_campos_basicos_cliente,
    validar_formulario_lead,
    validar_pergunta_segmento,
)
from pdf_builder import gerar_pdf, gerar_pdf_proposta_comercial
from utils import formatar_moeda
import io
import re
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload
import fitz
from openpyxl import load_workbook
from proposal_builder_v2 import gerar_pdf_proposta_comercial_v2
from proposal_html_builder import gerar_pdf_proposta_html

def autenticar_usuario(usuario, senha):
    try:
        for perfil, dados in st.secrets["auth"].items():
            if usuario == dados["username"] and senha == dados["password"]:
                return perfil
        return None
    except Exception:
        return None


def tela_login():
    st.title("🔐 Acesso Interno")
    st.write("Informe usuário e senha para acessar a área interna do CRM.")

    usuario = st.text_input("Usuário")
    senha = st.text_input("Senha", type="password")

    if st.button("Entrar"):
        perfil = autenticar_usuario(usuario, senha)

        if perfil:
            st.session_state["autenticado"] = True
            st.session_state["perfil_usuario"] = perfil
            st.success("Login realizado com sucesso.")
            st.rerun()
        else:
            st.error("Usuário ou senha inválidos.")

def estilo_status_linha(row):
    status = str(row.get("status_comercial", "")).strip()

    if status == "Em aberto":
        cor = "background-color: #fff3cd; color: #000000;"
    elif status == "Preço apresentado":
        cor = "background-color: #cfe2ff; color: #000000;"
    elif status == "Contrato fechado":
        cor = "background-color: #d1e7dd; color: #000000;"
    elif status == "Negativa":
        cor = "background-color: #f8d7da; color: #000000;"
    elif status == "Sem resposta":
        cor = "background-color: #e2e3e5; color: #000000;"
    else:
        cor = ""

    return [cor] * len(row)

def formatar_numero_br(valor):
    try:
        valor = float(valor or 0)
    except:
        valor = 0.0

    return f"{valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def converter_numero_br(texto):
    if texto is None:
        return 0.0

    texto = str(texto).strip()

    if texto == "":
        return 0.0

    texto = texto.replace("R$", "").strip()
    texto = texto.replace(".", "")
    texto = texto.replace(",", ".")

    try:
        return float(texto)
    except:
        return 0.0

def normalizar_regime_para_tabela(regime):
    if not regime:
        return ""
    regime = str(regime).strip()

    mapa = {
        "Simples": "Simples",
        "Simples Nacional": "Simples",
        "Presumido": "Lucro Presumido",
        "Lucro Presumido": "Lucro Presumido",
        "Real": "Lucro Real",
        "Lucro Real": "Lucro Real",
        "Não sei": "Simples",
    }
    return mapa.get(regime, regime)


def buscar_tabela_base(segmento):
    try:
        res = supabase.table("mapa_segmento_precificacao") \
            .select("tabela_base") \
            .eq("segmento_questionario", segmento) \
            .eq("ativo", True) \
            .limit(1) \
            .execute()

        if res.data:
            return res.data[0]["tabela_base"]
    except Exception as e:
        st.error(f"Erro ao buscar tabela base: {e}")

    return None


def buscar_preco_base_inicial(tabela_base, regime, faturamento):
    try:
        faturamento = float(faturamento or 0)
        regime_tabela = normalizar_regime_para_tabela(regime)

        res = supabase.table("precos_base_precificacao") \
            .select("*") \
            .eq("tabela_base", tabela_base) \
            .eq("regime", regime_tabela) \
            .eq("ativo", True) \
            .order("faixa_inicial") \
            .execute()

        if not res.data:
            return 0.0, None

        for linha in res.data:
            faixa_inicial = float(linha.get("faixa_inicial") or 0)
            faixa_final = float(linha.get("faixa_final") or 0)
            sem_limite_superior = bool(linha.get("sem_limite_superior"))
            valor_base = float(linha.get("valor_base") or 0)

            if sem_limite_superior:
                if faturamento >= faixa_inicial:
                    return valor_base, linha
            else:
                if faixa_inicial <= faturamento <= faixa_final:
                    return valor_base, linha

    except Exception as e:
        st.error(f"Erro ao buscar preço base inicial: {e}")

    return 0.0, None


def buscar_regras_precificacao(segmento_origem):
    try:
        res = supabase.table("regras_perguntas_precificacao") \
            .select("*") \
            .eq("segmento_origem", segmento_origem) \
            .eq("ativo", True) \
            .execute()

        return res.data if res.data else []
    except Exception as e:
        st.error(f"Erro ao buscar regras de precificação: {e}")
        return []

def limpar_nome_arquivo(texto):
    texto = str(texto or "").strip()
    texto = re.sub(r'[^A-Za-z0-9._ -]+', '', texto)
    texto = texto.replace(" ", "_")
    return texto or "sem_nome"


def get_drive_service():
    creds_info = dict(st.secrets["gcp_service_account"])
    scopes = ["https://www.googleapis.com/auth/drive.file"]

    credentials = Credentials.from_service_account_info(
        creds_info,
        scopes=scopes
    )

    return build("drive", "v3", credentials=credentials)

def arquivo_parece_balancete(uploaded_file):
    if not hasattr(uploaded_file, "name"):
        return False, "Nenhum arquivo válido foi enviado."

    nome = str(uploaded_file.name or "").lower()
    tipo = str(uploaded_file.type or "").lower()

    palavras_chave = [
        "balancete",
        "conta",
        "saldo",
        "saldo anterior",
        "saldo atual",
        "débito",
        "debito",
        "crédito",
        "credito",
        "ativo",
        "passivo",
        "receita",
        "despesa",
        "resultado",
    ]

    texto = ""

    try:
        if nome.endswith(".pdf") or "pdf" in tipo:
            pdf_bytes = uploaded_file.getvalue()
            doc = fitz.open(stream=pdf_bytes, filetype="pdf")

            for page in doc:
                texto += page.get_text("text") + "\n"

        elif nome.endswith(".xlsx") or nome.endswith(".xlsm"):
            arquivo = io.BytesIO(uploaded_file.getvalue())
            wb = load_workbook(arquivo, data_only=True, read_only=True)

            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    texto += " ".join([str(c) for c in row if c is not None]) + "\n"

        else:
            return False, "Formato não aceito. Envie PDF ou Excel."

        texto_lower = texto.lower()
        encontrados = [p for p in palavras_chave if p in texto_lower]
        
        obrigatorias = ["saldo", "conta"]
        
        if not all(p in texto_lower for p in obrigatorias):
            return False, "Arquivo não possui estrutura mínima de balancete. Precisa conter termos como saldo e conta."
        
        if len(encontrados) >= 5:
            return True, f"Arquivo validado como possível balancete. Termos encontrados: {', '.join(encontrados[:5])}"
        
        return False, "O arquivo não parece ser um balancete contábil. Verifique se enviou o documento correto."

    except Exception as e:
        return False, f"Não foi possível validar o arquivo: {e}"

def upload_arquivo_para_drive(uploaded_file, nome_empresa, lead_id, pasta_drive_id):
    service = get_drive_service()

    nome_empresa_limpo = limpar_nome_arquivo(nome_empresa)
    nome_original = limpar_nome_arquivo(uploaded_file.name)

    nome_salvo = f"{pd.Timestamp.today().date()}__lead_{lead_id}__{nome_empresa_limpo}__balancete__{nome_original}"

    file_metadata = {
        "name": nome_salvo,
        "parents": [pasta_drive_id]
    }

    file_bytes = io.BytesIO(uploaded_file.getvalue())

    media = MediaIoBaseUpload(
        file_bytes,
        mimetype=uploaded_file.type or "application/octet-stream",
        resumable=True
    )

    arquivo = service.files().create(
        body=file_metadata,
        media_body=media,
        fields="id, webViewLink",
        supportsAllDrives=True
    ).execute()

    return {
        "nome_original": uploaded_file.name,
        "nome_salvo": nome_salvo,
        "drive_file_id": arquivo.get("id"),
        "drive_link": arquivo.get("webViewLink"),
        "mime_type": uploaded_file.type
    }

# --- 1. CONFIGURAÇÃO DA PÁGINA ---
st.set_page_config(page_title="CRM & Precificação Escrita", layout="wide", page_icon="📄")

try:
    supabase = get_supabase()
except Exception as e:
    st.error(f"Erro de conexão com o Supabase: {e}")
    st.stop()

# --- 3. ESTILOS VISUAIS ---
st.markdown("""
    <style>
    .metric-card { 
        background-color: #1a2a44; padding: 25px; border-radius: 12px; 
        color: white; text-align: center; border: 1px solid #d4af37;
    }
    .metric-card h2 { color: #d4af37 !important; margin: 10px 0 !important; }
    div.stButton > button { border-radius: 5px; font-weight: bold; width: 100%; height: 3em; }
    </style>
    """, unsafe_allow_html=True)

def tela_lead_site():
    st.image("Logo Escrita.png", width=200)
    st.title("Fale com a Escrita Contabilidade")

    with st.form("form_site"):
        ajuda = st.text_area("Em que podemos ajudar? *", placeholder="Estou entrando em contato pois preciso...")
        nome = st.text_input("Nome Completo *")

        col1, col2 = st.columns(2)
        with col1:
            email = st.text_input("E-mail *")
        with col2:
            telefone = st.text_input("Telefone")

        possui_empresa = st.selectbox(
            "Você já possui uma empresa constituída? *",
            ["", "Sim", "Não"]
        )

        nome_empresa = st.text_input("Nome da Empresa")

        col3, col4 = st.columns(2)
        with col3:
            cnpj = st.text_input("CNPJ")
        with col4:
            cidade = st.text_input("Cidade")

        col5, col6, col7 = st.columns(3)
        with col5:
            uf = st.text_input("UF")
        with col6:
            faturamento_anual = st.text_input("Faturamento Anual")
        with col7:
            forma_tributacao = st.selectbox(
                "Forma de Tributação",
                ["", "Simples Nacional", "Lucro Presumido", "Lucro Real", "MEI", "Não sei"]
            )

        detalhes_empresa = st.text_area("Detalhes sobre a empresa")

        enviar = st.form_submit_button("Enviar")

        if enviar:
            if not ajuda or not nome or not email or not possui_empresa:
                st.warning("Preencha os campos obrigatórios marcados com *.")
                st.stop()

            obj = {
                "tipo_lead": "site",
                "origem": "formulario_site",
                "status": "Novo",
                "nome_empresa": nome_empresa,
                "responsavel": nome,
                "whatsapp": telefone,
                "email": email,
                "telefone": telefone,
                "possui_empresa_constituida": possui_empresa,
                "cnpj": cnpj,
                "cidade": cidade,
                "uf": uf,
                "faturamento_medio": faturamento_anual,
                "forma_tributacao": forma_tributacao,
                "descricao_atividades": ajuda,
                "detalhes_empresa": detalhes_empresa,
                "segmento": "Não informado",
                "regime": forma_tributacao,
                "respostas_segmento": {
                    "Em que podemos ajudar?": ajuda,
                    "Você já possui uma empresa constituída?": possui_empresa,
                    "Detalhes sobre a empresa": detalhes_empresa
                },
                "ativo": True
            }

            try:
                supabase.table("leads_externos").insert(obj).execute()
                st.success("Recebemos sua solicitação. Em breve entraremos em contato.")
                st.stop()
            except Exception as e:
                st.error(f"Erro ao salvar lead do site: {e}")
                
# --- 5. LÓGICA DE ACESSO (CLIENTE VS CONTADOR) ---
query_params = st.query_params
is_cliente = query_params.get("modo") == "cliente"
is_site = query_params.get("modo") == "site"

if is_site:
    tela_lead_site()
    st.stop()

if "autenticado" not in st.session_state:
    st.session_state["autenticado"] = False



if is_cliente:
    st.image("Logo Escrita.png", width=200)
    st.title("📝 Solicitação de Orçamento")
    st.write("Preencha os dados abaixo para receber nossa proposta comercial.")

    # 1. Busca os segmentos
    res_seg = supabase.table("segmentos").select("nome").execute()
    lista_segmentos = [s["nome"] for s in res_seg.data] if res_seg.data else []

    # 2. Cliente escolhe um ou mais segmentos
    res_regras = supabase.table("regras_segmento").select("segmentos").execute()
    lista_segmentos = [r["segmentos"] for r in res_regras.data] if res_regras.data else []
    
    f_segmento = st.selectbox(
        "Selecione o segmento da empresa",
        lista_segmentos
    )
    # 3. Busca perguntas pela origem correta
    res_perg_data = []

    if f_segmento:
        try:
            origem_perguntas = get_origem_perguntas(f_segmento)

            res_perg_data = get_perguntas_por_origem(origem_perguntas)
        except Exception as e:
            st.error(f"Erro ao carregar perguntas do segmento: {e}")

    with st.form("form_externo"):
        f_empresa = st.text_input("Nome da Empresa")
        f_cnpj = st.text_input("CNPJ")
        f_resp = st.text_input("Seu Nome")
        f_whatsapp = st.text_input("WhatsApp (com DDD)")
        f_regime = st.selectbox("Regime Atual", ["Simples", "Presumido", "Real", "Não sei"])
    
        st.divider()
        st.subheader("Informações Gerais")
    
        faturamento_medio_txt = st.text_input(
            "Faturamento médio mensal (R$)",
            value="0,00",
            placeholder="Ex: 1.000.000,00",
            key="cli_faturamento_txt"
        )
        
        faturamento_medio = converter_numero_br(faturamento_medio_txt)
        
        descricao_atividades = st.text_area(
            "Breve descrição sobre as atividades exercidas pela empresa",
            value="",
            key="cli_descricao"
        )
    
        respostas_extras = {}
        if res_perg_data:
            st.divider()
            st.subheader("Informações Adicionais")
    
            for p in res_perg_data:
                st.markdown(f"**{p['pergunta']}**")

                if "Múltipla Escolha" in p["tipo_campo"]:
                    ops = [o.strip() for o in str(p["opcoes"]).split(",") if o.strip()]
                    respostas_extras[p["pergunta"]] = st.radio(
                        "Selecione uma opção:",
                        ops,
                        key=f"ext_{p['id']}"
                    )
                elif p["tipo_campo"] == "Texto Livre":
                    pergunta_texto = str(p["pergunta"]).strip().lower()
                
                    if "enviar o último balancete" in pergunta_texto or "enviar o ultimo balancete" in pergunta_texto:
                        respostas_extras[p["pergunta"]] = st.file_uploader(
                            "Anexe o balancete",
                            type=["pdf", "xlsx", "xlsm"],
                            key=f"ext_balancete_{p['id']}"
                        )
                    else:
                        respostas_extras[p["pergunta"]] = st.text_area(
                            "Digite sua resposta:",
                            key=f"ext_{p['id']}"
                        )
                else:
                    respostas_extras[p["pergunta"]] = st.number_input(
                        "Informe a quantidade:",
                        min_value=0,
                        step=1,
                        key=f"ext_{p['id']}"
                    )

                st.write("")

        if st.form_submit_button("Enviar Solicitação"):
            erros = validar_formulario_lead(f_empresa, f_resp, f_whatsapp, f_segmento)
        
            if erros:
                for erro in erros:
                    st.warning(erro)
            else:
                try:
                    obj = {
                        "nome_empresa": f_empresa,
                        "cnpj": f_cnpj,
                        "responsavel": f_resp,
                        "whatsapp": f_whatsapp,
                        "regime": f_regime,
                        "segmento": f_segmento,
                        "faturamento_medio": faturamento_medio,
                        "descricao_atividades": descricao_atividades,
                        "respostas_segmento": {
                            k: (v.name if hasattr(v, "name") else v)
                            for k, v in respostas_extras.items()
                        }
                    }
        
                    cnpj_limpo = "".join([c for c in str(f_cnpj or "") if c.isdigit()])

                    if cnpj_limpo:
                        obj["cnpj"] = cnpj_limpo
                    
                        res_existente = supabase.table("leads_externos") \
                            .select("*") \
                            .eq("cnpj", cnpj_limpo) \
                            .eq("ativo", True) \
                            .limit(1) \
                            .execute()
                    
                        if res_existente.data:
                            lead_id_existente = res_existente.data[0]["id"]
                    
                            res_insert = supabase.table("leads_externos") \
                                .update(obj) \
                                .eq("id", lead_id_existente) \
                                .execute()
                        else:
                            res_insert = supabase.table("leads_externos").insert(obj).execute()
                    else:
                        res_insert = supabase.table("leads_externos").insert(obj).execute()
        
                    if not res_insert.data:
                        st.error("Não foi possível salvar o lead.")
                        st.stop()
        
                    lead_salvo = res_insert.data[0]
                    lead_id = lead_salvo["id"]
        
                    for pergunta, valor in respostas_extras.items():
                        pergunta_texto = str(pergunta).strip().lower()
        
                        if "balancete" in pergunta_texto and hasattr(valor, "name"):
                            arquivo_ok, mensagem_validacao = arquivo_parece_balancete(valor)

                            if not arquivo_ok:
                                st.error(mensagem_validacao)
                                st.stop()
                                
                            pasta_drive_id = st.secrets["drive_balancetes_folder_id"]
        
                            arquivo_info = upload_arquivo_para_drive(
                                uploaded_file=valor,
                                nome_empresa=f_empresa,
                                lead_id=lead_id,
                                pasta_drive_id=pasta_drive_id
                            )
        
                            supabase.table("lead_arquivos").insert({
                                "lead_id": lead_id,
                                "tipo_arquivo": "balancete",
                                "nome_original": arquivo_info["nome_original"],
                                "nome_salvo": arquivo_info["nome_salvo"],
                                "drive_file_id": arquivo_info["drive_file_id"],
                                "drive_link": arquivo_info["drive_link"],
                                "mime_type": arquivo_info["mime_type"]
                            }).execute()
        
                    st.success("✅ Recebemos seus dados e o balancete foi anexado com sucesso.")
                    st.stop()
        
                except Exception as e:
                    st.error(f"Erro ao salvar lead/anexar balancete: {e}")
else:
    if not st.session_state["autenticado"]:
        tela_login()
        st.stop()
        
    if os.path.exists("Logo Escrita.png"):
        st.sidebar.image("Logo Escrita.png", width=200)

    if st.sidebar.button("Sair"):
        st.session_state["autenticado"] = False
        st.session_state["perfil_usuario"] = ""
        st.rerun()
        
    perfil_usuario = st.session_state.get("perfil_usuario", "")

    menus_permitidos = [
        "Leads Recebidos",
        "Leads Arquivados",
        "Nova Proposta",
        "Proposta Comercial",
        "Histórico de Vendas",
        "Link para Cliente"
    ]
    
    if perfil_usuario == "admin":
        menus_permitidos.insert(3, "Dashboard de Custos")
        menus_permitidos.append("Configurações")
    
    menu = st.sidebar.selectbox(
        "Navegação",
        menus_permitidos
    )

    if menu == "Leads Recebidos":
        st.title("📥 Leads Recebidos")

        try:
            res_leads = supabase.table("leads_externos") \
                .select("*") \
                .eq("ativo", True) \
                .order("created_at", desc=True) \
                .execute()

            if res_leads.data:
                df_leads = pd.DataFrame(res_leads.data)

                colunas_exibir = [
                    "id",
                    "nome_empresa",
                    "responsavel",
                    "segmento",
                    "regime",
                    "status",
                    "created_at"
                ]
                colunas_exibir = [c for c in colunas_exibir if c in df_leads.columns]

                st.dataframe(df_leads[colunas_exibir], use_container_width=True)

                lista_opcoes = [
                    f"{row['id']} | {row['nome_empresa']} | {row['segmento']}"
                    for _, row in df_leads.iterrows()
                ]

                lead_escolhido = st.selectbox(
                    "Selecione um lead para analisar",
                    lista_opcoes
                )

                if st.button("Carregar para Precificação"):
                    lead_id = int(lead_escolhido.split("|")[0].strip())
                    lead_data = df_leads[df_leads["id"] == lead_id].iloc[0].to_dict()

                    st.session_state["lead_em_analise"] = lead_data

                    try:
                        supabase.table("leads_externos").update({
                            "status": "Em análise"
                        }).eq("id", lead_id).execute()
                    except Exception as e:
                        st.warning(f"Não foi possível atualizar o status do lead: {e}")

                    st.success("Lead carregado. Agora vá para 'Nova Proposta'.")
                    
                if st.button("Arquivar Lead Selecionado"):
                    lead_id = int(lead_escolhido.split("|")[0].strip())
                
                    try:
                        supabase.table("leads_externos").update({
                            "ativo": False,
                            "deleted_at": pd.Timestamp.now().isoformat(),
                            "deleted_reason": "Arquivado manualmente no CRM"
                        }).eq("id", lead_id).execute()
                
                        st.success("Lead arquivado com sucesso.")
                        st.rerun()
                
                    except Exception as e:
                        st.error(f"Erro ao arquivar lead: {e}")    

                    
            else:
                st.info("Nenhum lead recebido ainda.")

        except Exception as e:
            st.error(f"Erro ao carregar leads: {e}")
            
    elif menu == "Leads Arquivados":
        st.title("📦 Leads Arquivados")

        try:
            res_leads = supabase.table("leads_externos") \
                .select("*") \
                .eq("ativo", False) \
                .order("deleted_at", desc=True) \
                .execute()

            if not res_leads.data:
                st.info("Nenhum lead arquivado.")
            else:
                df_leads = pd.DataFrame(res_leads.data)

                colunas_exibir = [
                    "id",
                    "nome_empresa",
                    "responsavel",
                    "cnpj",
                    "segmento",
                    "regime",
                    "status",
                    "deleted_at",
                    "deleted_reason"
                ]
                colunas_exibir = [c for c in colunas_exibir if c in df_leads.columns]

                st.dataframe(df_leads[colunas_exibir], use_container_width=True)

                lista_opcoes = [
                    f"{row['id']} | {row.get('nome_empresa', 'Sem nome')} | {row.get('responsavel', '')}"
                    for _, row in df_leads.iterrows()
                ]

                lead_escolhido = st.selectbox(
                    "Selecione um lead para restaurar",
                    lista_opcoes
                )

                if st.button("Restaurar Lead Selecionado"):
                    lead_id = int(lead_escolhido.split("|")[0].strip())

                    supabase.table("leads_externos").update({
                        "ativo": True,
                        "deleted_at": None,
                        "deleted_reason": None
                    }).eq("id", lead_id).execute()

                    st.success("Lead restaurado com sucesso.")
                    st.rerun()

        except Exception as e:
            st.error(f"Erro ao carregar leads arquivados: {e}")
            
    if menu == "Nova Proposta":
            st.title("📄 Elaboração de Proposta Precificada")
            lead_em_analise = st.session_state.get("lead_em_analise", {})
        
            # 1. Inputs de Identificação e Regime
            c1, c2 = st.columns([2, 1])

            nome_cliente = c1.text_input(
                "Nome da Empresa:",
                value=lead_em_analise.get("nome_empresa", "")
            )
            
            opcoes_regime = ["Simples", "Presumido", "Real"]
            regime_padrao = lead_em_analise.get("regime", "Simples")
            if regime_padrao not in opcoes_regime:
                regime_padrao = "Simples"
            
            regime_sel = c2.selectbox(
                "Regime Tributário:",
                opcoes_regime,
                index=opcoes_regime.index(regime_padrao)
            )

            st.divider()
            st.subheader("Informações Gerais")
            
            faturamento_medio = st.number_input(
                "Faturamento médio mensal (R$)",
                min_value=0.0,
                step=1000.0,
                format="%.2f",
                value=float(lead_em_analise.get("faturamento_medio") or 0),
                key="np_faturamento"
            )
            
            descricao_atividades = st.text_area(
                "Breve descrição sobre as atividades exercidas pela empresa",
                value=lead_em_analise.get("descricao_atividades", "") or "",
                key="cli_descricao"
            )
    
            # 2. Seleção de Segmento para carregar as Perguntas
            res_seg = supabase.table("segmentos").select("*").execute()
            lista_s = [s["nome"] for s in res_seg.data] if res_seg.data else []
            
            res_regras = supabase.table("regras_segmento").select("segmentos").execute()
            lista_segmentos = [r["segmentos"] for r in res_regras.data] if res_regras.data else []
    
            segmento_padrao = lead_em_analise.get("segmento", "")
            if segmento_padrao not in lista_segmentos and lista_segmentos:
                segmento_padrao = lista_segmentos[0]
    
            seg_sel = st.selectbox(
                "Selecione o segmento do cliente:",
                lista_segmentos,
                index=lista_segmentos.index(segmento_padrao) if segmento_padrao in lista_segmentos else 0
            )
    
            st.divider()

            # 4. Perguntas Dinâmicas do Segmento + Novo Motor de Precificação
            respostas_lead = lead_em_analise.get("respostas_segmento", {}) or {}
            if not isinstance(respostas_lead, dict):
                respostas_lead = {}

            respostas_formulario = {}
            res_perg_data = []

            if seg_sel:
                try:
                    origem_perguntas = get_origem_perguntas(seg_sel)
                    res_perg_data = get_perguntas_por_origem(origem_perguntas)
                    st.caption(f"Origem das perguntas utilizada: {origem_perguntas}")
                except Exception as e:
                    st.error(f"Erro ao carregar perguntas do segmento: {e}")
                    origem_perguntas = seg_sel
            else:
                origem_perguntas = ""

            if res_perg_data:
                st.subheader(f"📋 Diagnóstico Específico: {seg_sel}")

                for p in res_perg_data:
                    st.markdown(f"**{p['pergunta']}**")

                    pergunta_texto = str(p.get("pergunta", "")).strip()
                    resposta_inicial = respostas_lead.get(pergunta_texto, None)

                    if "Múltipla Escolha" in p["tipo_campo"]:
                        ops = [o.strip() for o in str(p.get("opcoes", "")).split(",") if o.strip()]

                        indice_padrao = 0
                        if resposta_inicial in ops:
                            indice_padrao = ops.index(resposta_inicial)

                        resposta = st.radio(
                            "Selecione uma opção:",
                            ops,
                            index=indice_padrao,
                            key=f"p_{p['id']}"
                        )

                    elif p["tipo_campo"] == "Texto Livre":
                        resposta = st.text_area(
                            "Digite sua resposta:",
                            value=str(resposta_inicial or ""),
                            key=f"p_{p['id']}"
                        )

                    else:
                        valor_inicial = 0
                        try:
                            if resposta_inicial not in [None, ""]:
                                valor_inicial = int(float(resposta_inicial))
                        except:
                            valor_inicial = 0

                        resposta = st.number_input(
                            "Informe a quantidade:",
                            min_value=0,
                            step=1,
                            value=valor_inicial,
                            key=f"p_{p['id']}"
                        )

                    respostas_formulario[pergunta_texto] = resposta
                    st.write("")

            st.divider()

            # 5. Validação e novo cálculo
            erros = validar_campos_basicos_cliente(nome_cliente, regime_sel, seg_sel)
            for erro in erros:
                st.warning(erro)

            valores = None
            memoria = None

            if not erros:
                try:
                    tabela_base = buscar_tabela_base(seg_sel)
                    preco_base_inicial, faixa_encontrada = buscar_preco_base_inicial(
                        tabela_base=tabela_base,
                        regime=regime_sel,
                        faturamento=faturamento_medio
                    )

                    regras_precificacao = buscar_regras_precificacao(origem_perguntas)
                    

                    preco_base_calculado, total_acrescimos, detalhamento_acrescimos = calcular_preco_completo(
                        valor_base=preco_base_inicial,
                        respostas_formulario=respostas_formulario,
                        regras=regras_precificacao
                    )
                    
                    v_bronze = preco_base_calculado 
                    v_prata = preco_base_calculado * 1.15
                    v_ouro = preco_base_calculado * 1.35

                    valores = {
                        "bronze": v_bronze,
                        "prata": v_prata,
                        "ouro": v_ouro
                    }

                    memoria = {
                        "segmento_escolhido": seg_sel,
                        "origem_perguntas": origem_perguntas,
                        "tabela_base": tabela_base,
                        "regime": regime_sel,
                        "faturamento_medio": faturamento_medio,
                        "faixa_encontrada": faixa_encontrada,
                        "preco_base_inicial": preco_base_inicial,
                        "total_acrescimos": total_acrescimos,
                        "detalhamento_acrescimos": detalhamento_acrescimos,
                        "preco_base_calculado": preco_base_calculado,
                        "valor_bronze": v_bronze,
                        "valor_prata": v_prata,
                        "valor_ouro": v_ouro
                    }

                except Exception as e:
                    st.error(f"Erro no cálculo da precificação: {e}")

            # 6. Exibição dos cards
            if valores:
                st.subheader("💰 Opções de Investimento")
                res1, res2, res3 = st.columns(3)

                v_bronze = valores["bronze"]
                v_prata = valores["prata"]
                v_ouro = valores["ouro"]

                res1.markdown(
                    f"""<div class="metric-card"><p>BRONZE </p><h2>{formatar_moeda(v_bronze)}</h2></div>""",
                    unsafe_allow_html=True
                )
                res2.markdown(
                    f"""<div class="metric-card"><p>PRATA </p><h2>{formatar_moeda(v_prata)}</h2></div>""",
                    unsafe_allow_html=True
                )
                res3.markdown(
                    f"""<div class="metric-card"><p>OURO </p><h2>{formatar_moeda(v_ouro)}</h2></div>""",
                    unsafe_allow_html=True
                )

                with st.expander("Ver memória do cálculo"):
                    st.json(memoria)

                st.session_state["proposta_atual"] = {
                    "cliente": nome_cliente,
                    "regime": regime_sel,
                    "segmento": seg_sel,
                    "origem_perguntas": origem_perguntas,
                    "tabela_base": memoria["tabela_base"],
                    "faturamento_medio": faturamento_medio,
                    "descricao_atividades": descricao_atividades,
                    "respostas_formulario": respostas_formulario,
                    "preco_base_inicial": memoria["preco_base_inicial"],
                    "total_acrescimos": memoria["total_acrescimos"],
                    "preco_base_calculado": memoria["preco_base_calculado"],
                    "valor_bronze": v_bronze,
                    "valor_prata": v_prata,
                    "valor_ouro": v_ouro,
                    "valor_escolhido": v_prata
                }

                # 7. Salvamento
                if st.button("💾 Salvar Orçamento Final"):
                    try:
                        lead_id_atual = lead_em_analise.get("id")

                        dados_venda = {
                            "lead_id": lead_id_atual,
                            "cliente": nome_cliente,
                            "regime": regime_sel,
                            "segmento": seg_sel,
                            "faturamento_medio": faturamento_medio,
                            "descricao_atividades": descricao_atividades,
                            "valor_total": v_prata,
                            "observacoes_comerciais": str(memoria)
                        }
                        insert_data("historico_vendas", dados_venda)
                        st.success("Orçamento salvo com sucesso!")
                    except Exception as e:
                        st.error(f"Erro ao salvar orçamento: {e}")

    
    # --- MÓDULOS DE APOIO (MANTIDOS E INTEGRADOS) ---
    
    elif menu == "Proposta Comercial":
        st.title("📑 Proposta Comercial")

        proposta_atual = st.session_state.get("proposta_atual", {})

        nome_empresa = proposta_atual.get("cliente", "")
        segmento = proposta_atual.get("segmento", "")
        regime = proposta_atual.get("regime", "")
        tabela_base = proposta_atual.get("tabela_base", "")
        faturamento_medio = proposta_atual.get("faturamento_medio", 0.0)
        preco_base_inicial = proposta_atual.get("preco_base_inicial", 0.0)
        total_acrescimos = proposta_atual.get("total_acrescimos", 0.0)
        preco_base_calculado = proposta_atual.get("preco_base_calculado", 0.0)

        valor_bronze = proposta_atual.get("valor_bronze", 0.0)
        valor_prata = proposta_atual.get("valor_prata", 0.0)
        valor_ouro = proposta_atual.get("valor_ouro", 0.0)

        if not nome_empresa:
            st.warning("Gere uma proposta primeiro na tela 'Nova Proposta'.")
        else:
            st.subheader("Resumo da Proposta")
            c1, c2 = st.columns(2)

            with c1:
                st.text_input("Empresa", value=nome_empresa, disabled=True)
                st.text_input("Segmento", value=segmento, disabled=True)
                st.text_input("Regime", value=regime, disabled=True)

            with c2:
                st.text_input("Tabela Base Utilizada", value=tabela_base, disabled=True)
                st.text_input("Faturamento Médio", value=formatar_moeda(faturamento_medio), disabled=True)

            st.divider()

            st.subheader("Formação do Preço")
            c3, c4, c5 = st.columns(3)

            with c3:
                st.text_input("Preço Base Inicial", value=formatar_moeda(preco_base_inicial), disabled=True)

            with c4:
                st.text_input("Acréscimos do Questionário", value=formatar_moeda(total_acrescimos), disabled=True)

            with c5:
                st.text_input("Preço Base Calculado", value=formatar_moeda(preco_base_calculado), disabled=True)

            st.divider()

            st.subheader("Plano a Apresentar")
            opcao_valor = st.selectbox(
                "Selecione o plano comercial",
                ["Bronze", "Prata", "Ouro"],
                index=1
            )

            if opcao_valor == "Bronze":
                valor_apresentado = valor_bronze
            elif opcao_valor == "Ouro":
                valor_apresentado = valor_ouro
            else:
                valor_apresentado = valor_prata

            proposta_atual["plano_escolhido"] = opcao_valor
            proposta_atual["valor_escolhido"] = valor_apresentado
            st.session_state["proposta_atual"] = proposta_atual

            c6, c7, c8 = st.columns(3)

            with c6:
                st.markdown(
                    f"""<div class="metric-card"><p>BRONZE</p><h2>{formatar_moeda(valor_bronze)}</h2></div>""",
                    unsafe_allow_html=True
                )

            with c7:
                st.markdown(
                    f"""<div class="metric-card"><p>PRATA</p><h2>{formatar_moeda(valor_prata)}</h2></div>""",
                    unsafe_allow_html=True
                )

            with c8:
                st.markdown(
                    f"""<div class="metric-card"><p>OURO</p><h2>{formatar_moeda(valor_ouro)}</h2></div>""",
                    unsafe_allow_html=True
                )

            st.info(f"Plano selecionado para apresentação: {opcao_valor} — {formatar_moeda(valor_apresentado)}")

            st.subheader("Serviços contratados")

            servicos_contratados = st.multiselect(
                "Selecione os serviços incluídos na proposta:",
                ["Contábil", "Fiscal", "Pessoal", "Societário"],
                default=["Contábil", "Fiscal", "Pessoal", "Societário"]
            )
            
            if not servicos_contratados:
                st.warning("Selecione pelo menos um serviço para gerar a proposta.")
                
            with st.expander("Ver detalhes do cálculo da proposta"):
                st.json(proposta_atual)

            st.divider()

            
            st.divider()

            if st.button("📄 Preparar PDF Profissional"):
                if not servicos_contratados:
                    st.error("Selecione pelo menos um serviço contratado antes de gerar a proposta.")
                    st.stop()
            
                try:
                    caminho_pdf = gerar_pdf_proposta_html(
                        nome_empresa=nome_empresa,
                        plano=opcao_valor,
                        valor_mensal=valor_apresentado,
                        servicos_contratados=servicos_contratados,
                        respostas_cliente=proposta_atual.get("respostas_formulario", {})
                    )
            
                    st.session_state["pdf_proposta_path"] = caminho_pdf
                    st.success("PDF profissional preparado com sucesso.")
            
                except Exception as e:
                    st.error(f"Erro ao preparar PDF profissional: {e}")
                    
            caminho_pdf = st.session_state.get("pdf_proposta_path")

            if caminho_pdf and os.path.exists(caminho_pdf):
                with open(caminho_pdf, "rb") as arquivo_pdf:
                    st.download_button(
                        label="⬇️ Baixar PDF da Proposta",
                        data=arquivo_pdf,
                        file_name=f"proposta_{nome_empresa.replace(' ', '_')}.pdf",
                        mime="application/pdf"
                    )
    
    elif menu == "Dashboard de Custos":
        st.title("💰 Configuração de Custos Operacionais")
        
        col1, col2 = st.columns(2)
        with col1:
            st.subheader("Custos de Estrutura")
            f_folha = st.number_input("Total Folha + Encargos (R$)", value=get_config_val('total_folha'))
            f_fixas = st.number_input("Despesas Fixas (Sistemas/Aluguel) (R$)", value=get_config_val('despesas_fixas'))
            f_imposto = st.number_input("Imposto Médio s/ Faturamento (%)", value=get_config_val('impostos_faturamento'))
        
        with col2:
            st.subheader("Capacidade Produtiva")
            f_horas = st.number_input("Horas Úteis por Colaborador/Mês", value=get_config_val('horas_uteis_mes'))
            f_equipe = st.number_input("Quantidade de Colaboradores", value=get_config_val('num_colaboradores'))
    
        if st.button("💾 Salvar e Atualizar Custo-Hora"):
            configs = [
                {"chave": "total_folha", "valor": f_folha},
                {"chave": "despesas_fixas", "valor": f_fixas},
                {"chave": "impostos_faturamento", "valor": f_imposto},
                {"chave": "horas_uteis_mes", "valor": f_horas},
                {"chave": "num_colaboradores", "valor": f_equipe}
            ]
            for c in configs:
                supabase.table("configuracao_operacional").upsert(c, on_conflict="chave").execute()
            st.success("Custo-Hora atualizado!")
            st.rerun()
    
        c_hora = calcular_custo_hora_real()
        st.divider()
        st.markdown(f"""<div class="metric-card"><p>Custo Hora Atual</p><h2>{formatar_moeda(c_hora)}</h2></div>""", unsafe_allow_html=True)
    
    elif menu == "Link para Cliente":
        st.title("🔗 Coleta Externa de Dados")
        st.info("Envie o link abaixo para o prospecto preencher as informações iniciais.")
        st.code("https://crm-escrita-contabilidade.streamlit.app/?modo=cliente&embed=true")
        
        st.divider()
        st.subheader("📥 Leads Recebidos")
        try:
            res = supabase.table("leads_externos").select("*").order("created_at", desc=True).execute()
            if res.data:
                st.dataframe(pd.DataFrame(res.data), use_container_width=True)
            else:
                st.write("Nenhum lead preenchido ainda.")
        except Exception as e:
            st.error(f"Erro ao carregar leads: {e}")
            
    elif menu == "Histórico de Vendas":
        st.title("📊 Histórico de Orçamentos")

        try:
            res_h = supabase.table("historico_vendas").select("*").order("data_criacao", desc=True).execute()

            if not res_h.data:
                st.info("Nenhum orçamento salvo ainda.")
            else:
                df_h = pd.DataFrame(res_h.data)

                if "status_comercial" not in df_h.columns:
                    df_h["status_comercial"] = "Em aberto"

                filtro_status = st.selectbox(
                    "Filtrar por status",
                    ["Todos", "Em aberto", "Preço apresentado", "Contrato fechado", "Negativa", "Sem resposta"]
                )

                if filtro_status != "Todos":
                    df_h = df_h[df_h["status_comercial"] == filtro_status]

                st.subheader("Atualizar status comercial")

                opcoes_linha = [
                    f"{row['id']} | {row.get('cliente', 'Sem nome')} | {formatar_moeda(row.get('valor_total', 0))}"
                    for _, row in df_h.iterrows()
                ]

                if opcoes_linha:
                    linha_escolhida = st.selectbox("Selecione um orçamento", opcoes_linha)

                    novo_status = st.selectbox(
                        "Novo status",
                        ["Em aberto", "Preço apresentado", "Contrato fechado", "Negativa", "Sem resposta"]
                    )

                    observacao_status = st.text_input("Observação do status", value="")

                    if st.button("Salvar status comercial"):
                        try:
                            id_escolhido = int(linha_escolhida.split("|")[0].strip())

                            dados_update = {
                                "status_comercial": novo_status,
                                "observacao_status": observacao_status
                            }

                            if novo_status == "Preço apresentado":
                                dados_update["data_apresentacao"] = pd.Timestamp.today().date().isoformat()

                            elif novo_status == "Contrato fechado":
                                dados_update["data_fechamento"] = pd.Timestamp.today().date().isoformat()

                            elif novo_status == "Negativa":
                                dados_update["data_negativa"] = pd.Timestamp.today().date().isoformat()

                            supabase.table("historico_vendas").update(dados_update).eq("id", id_escolhido).execute()
                            # atualiza o lead vinculado, se existir
                            try:
                                lead_id_vinculado = df_h[df_h["id"] == id_escolhido]["lead_id"].iloc[0] if "lead_id" in df_h.columns else None

                                if pd.notnull(lead_id_vinculado):
                                    status_lead = "Em análise"

                                    if novo_status == "Preço apresentado":
                                        status_lead = "Preço apresentado"
                                    elif novo_status == "Contrato fechado":
                                        status_lead = "Fechado"
                                    elif novo_status == "Negativa":
                                        status_lead = "Negativa"
                                    elif novo_status == "Sem resposta":
                                        status_lead = "Sem resposta"
                                    elif novo_status == "Em aberto":
                                        status_lead = "Em análise"

                                    supabase.table("leads_externos").update({
                                        "status": status_lead
                                    }).eq("id", int(lead_id_vinculado)).execute()

                            except Exception as e:
                                st.warning(f"Status do orçamento salvo, mas não foi possível atualizar o lead: {e}")
                            st.success("Status atualizado com sucesso.")
                            st.rerun()

                        except Exception as e:
                            st.error(f"Erro ao atualizar status: {e}")

                st.divider()

                # cálculo de dias
                hoje = pd.Timestamp.today().normalize()

                if "data_apresentacao" in df_h.columns:
                    df_h["dias_desde_apresentacao"] = df_h["data_apresentacao"].apply(
                        lambda x: (hoje - pd.to_datetime(x)).days if pd.notnull(x) and x != "" else None
                    )
                else:
                    df_h["dias_desde_apresentacao"] = None

                if "data_fechamento" in df_h.columns:
                    df_h["dias_ate_fechamento"] = df_h.apply(
                        lambda row: (
                            (pd.to_datetime(row["data_fechamento"]) - pd.to_datetime(row["data_apresentacao"])).days
                            if pd.notnull(row.get("data_fechamento")) and pd.notnull(row.get("data_apresentacao"))
                            and row.get("data_fechamento") != "" and row.get("data_apresentacao") != ""
                            else None
                        ),
                        axis=1
                    )
                else:
                    df_h["dias_ate_fechamento"] = None

                colunas_preferidas = [
                    "id",
                    "cliente",
                    "segmento",
                    "regime",
                    "valor_total",
                    "status_comercial",
                    "data_apresentacao",
                    "data_fechamento",
                    "data_negativa",
                    "dias_desde_apresentacao",
                    "dias_ate_fechamento",
                    "observacao_status",
                    "data_criacao",
                ]

                colunas_exibir = [c for c in colunas_preferidas if c in df_h.columns]

                if "valor_total" in df_h.columns:
                    df_h["valor_total"] = df_h["valor_total"].apply(formatar_moeda)

                st.subheader("Histórico")

                st.dataframe(
                    df_h[colunas_exibir].style.apply(estilo_status_linha, axis=1),
                    use_container_width=True
                )

        except Exception as e:
            st.error(f"Erro ao carregar histórico de vendas: {e}")
    
    elif menu == "Configurações":
        st.title("⚙️ Painel de Controle e Cadastros")
        t1, t2, t3, t4 = st.tabs(["Segmentos e Perguntas", "Preços Avulsos", "Pesos de Esforço", "Custos Fixos"])
        
        with t1:
            st.subheader("1. Gestão de Segmentos")
            col_seg1, col_seg2 = st.columns([1, 2])
            with col_seg1:
                n_seg = st.text_input("Novo Segmento (Ex: Clínica):")
                if st.button("Salvar Segmento"):
                    supabase.table("segmentos").insert({"nome": n_seg}).execute()
                    st.rerun()
            with col_seg2:
                res_s = supabase.table("segmentos").select("*").execute()
                if res_s.data:
                    df_s = pd.DataFrame(res_s.data)
                    st.write("Segmentos Cadastrados:")
                    st.dataframe(df_s[['nome']], use_container_width=True)
    
            st.divider()
            st.subheader("2. Gestão de Perguntas")
            with st.form("nova_pergunta"):
                f_seg = st.selectbox("Segmento Alvo", [s['nome'] for s in res_s.data] if res_s.data else [])
                f_tipo = st.selectbox("Tipo", ["Múltipla Escolha", "Número (Multiplicador)"])
                f_perg = st.text_input("Pergunta")
                f_opt = st.text_input("Opções (Ex: Sim, Não ou Pequeno, Médio)")
                f_pesos = st.text_input("Pesos (Ex: 100, 0 ou 50, 150)")
                if st.form_submit_button("Salvar Pergunta"):
                    erros = validar_pergunta_segmento(f_tipo, f_perg, f_opt, f_pesos)

                    if erros:
                        for erro in erros:
                            st.warning(erro)
                    else:
                        try:
                            supabase.table("perguntas").insert({
                                "segmento": f_seg,
                                "pergunta": f_perg,
                                "tipo_campo": f_tipo,
                                "opcoes": f_opt,
                                "pesos_opcoes": f_pesos
                            }).execute()
                            st.success("Pergunta salva!")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Erro ao salvar pergunta: {e}")
    
            # Visualização das Perguntas com Filtro
            st.write("---")
            st.write("🔍 Perguntas Existentes")
            filtro_p = st.selectbox("Filtrar por Segmento:", ["Todos"] + ([s['nome'] for s in res_s.data] if res_s.data else []))
            query_p = supabase.table("perguntas").select("*")
            if filtro_p != "Todos":
                query_p = query_p.eq("segmento", filtro_p)
            res_p = query_p.execute()
            if res_p.data:
                df_perg = pd.DataFrame(res_p.data)
            
                colunas_desejadas = ['origem', 'pergunta', 'tipo_campo', 'opcoes', 'pesos_opcoes']
                colunas_existentes = [c for c in colunas_desejadas if c in df_perg.columns]
            
                st.dataframe(df_perg[colunas_existentes], use_container_width=True)
            else:
                st.info("Nenhuma pergunta cadastrada.")
                   
        with t2:
            st.subheader("Serviços Avulsos (Tabela 2026)")
            with st.form("add_avulso"):
                st.text_input("Nome do Serviço", key="av_n")
                st.number_input("Valor (R$)", key="av_v")
                if st.form_submit_button("Adicionar Serviço"):
                    supabase.table("servicos_avulsos").insert({"servico": st.session_state.av_n, "valor": st.session_state.av_v}).execute()
                    st.rerun()
            
            res_av = supabase.table("servicos_avulsos").select("*").execute()
            if res_av.data:
                st.write("Lista de Serviços Cadastrados:")
                st.dataframe(pd.DataFrame(res_av.data)[['servico', 'valor']], use_container_width=True)
    
        with t3:
            st.subheader("Ajuste de Pesos de Esforço (Horas)")
            st.info("Aqui você altera quanto tempo cada item (Nota, Funcionário, etc) consome em cada regime.")
            res_pesos = supabase.table("pesos_esforco").select("*").execute()
            if res_pesos.data:
                df_pesos = pd.DataFrame(res_pesos.data)
                # Filtro por Regime
                reg_f = st.selectbox("Filtrar Regime:", ["Todos", "Simples", "Presumido", "Real", "Filial"])
                df_p_view = df_pesos if reg_f == "Todos" else df_pesos[df_pesos['regime'] == reg_f]
                st.dataframe(df_p_view[['regime', 'item', 'horas_esforco']], use_container_width=True)
                
                st.warning("Para editar esses valores, utilize o Table Editor do Supabase diretamente por enquanto (Segurança do Banco).")
    
        with t4:
            st.subheader("Custos Fixos (Visualização)")
            res_cf = supabase.table("configuracao_operacional").select("*").execute()
            if res_cf.data:
                st.dataframe(pd.DataFrame(res_cf.data)[['chave', 'valor']], use_container_width=True)
